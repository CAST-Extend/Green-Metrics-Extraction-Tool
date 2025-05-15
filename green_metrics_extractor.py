import requests
import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, numbers
from datetime import datetime
import os

def load_config():
    try:
        with open('config.json', 'r') as f:
            config = json.load(f)
            return (
                config.get('HLInstance'),
                config.get('domain_id'),
                config.get('application_id'),
                config.get('api_key')
            )
    except Exception as e:
        print(f"Error loading config.json: {e}")
        return None, None, None, None

def get_api_data(hl_instance, domain_id, application_id, api_key):
    url = f'https://{hl_instance}.casthighlight.com/WS2/domains/{domain_id}/applications/{application_id}'
    headers = {'Authorization': f'Bearer {api_key}'}
    try:
        response = requests.get(url, headers=headers, timeout=30)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"API request failed: {e}")
        return None

def extract_green_data(hl_instance, domain_id, application_id, api_key):
    data = get_api_data(hl_instance, domain_id, application_id, api_key)
    if not data:
        return None, None

    if 'metrics' not in data or not data['metrics']:
        print("No metrics data found in the response.")
        return None, None

    metric = data['metrics'][0]
    if 'greenDetail' not in metric or not metric['greenDetail']:
        print("No green details found in metrics.")
        return None, None

    rows = []
    for tech_data in metric['greenDetail']:
        technology = tech_data.get('technology', 'N/A')
        
        if 'greenIndexDetails' not in tech_data or not tech_data['greenIndexDetails']:
            continue
            
        for detail in tech_data['greenIndexDetails']:
            green_req = detail.get('greenRequirement', {})
            occurrences = detail.get('greenOccurrences', 0)
            
            if occurrences == 0:
                continue
                
            effort_min = detail.get('greenEffort', 0)
            effort_person_days = round(effort_min / 480, 2)  # 1 day = 480 minutes
            
            rows.append([
                green_req.get('display', 'N/A'),
                technology,
                occurrences,
                effort_person_days,
                None,  # Placeholder for cost input
                None   # Placeholder for Tech Debt calculation
            ])
    
    if not rows:
        print("No rules with occurrences found.")
        return None, None
        
    # Create detailed DataFrame
    detailed_df = pd.DataFrame(rows, columns=[
        'Rule/Pattern',
        'Technology',
        'Number of Occurrences',
        'Effort by Occurrence (Person-day)',
        'Cost (FTE/Day)',
        'Tech Debt ($) Effort x Cost'
    ]).sort_values('Number of Occurrences', ascending=False)
    
    # Create summary DataFrame
    summary_data = []
    for rule, group in detailed_df.groupby('Rule/Pattern'):
        technologies = ', '.join(group['Technology'].unique())
        total_occurrences = group['Number of Occurrences'].sum()
        summary_data.append([rule, technologies, total_occurrences])
    
    summary_df = pd.DataFrame(summary_data, columns=[
        'Rule/Pattern',
        'Technology',
        'Number of Occurrences'
    ]).sort_values('Number of Occurrences', ascending=False)
    
    return detailed_df, summary_df

def save_to_excel(detailed_df, summary_df, domain_id, application_id):
    os.makedirs('output', exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f'output/green_metrics_d{domain_id}_a{application_id}_{timestamp}.xlsx'
    
    # Create a copy of the DataFrames for Excel export
    excel_detailed_df = detailed_df.copy()
    excel_summary_df = summary_df.copy()
    
    # Add total row to detailed sheet
    total_row = {
        'Rule/Pattern': 'TOTAL',
        'Technology': '',
        'Number of Occurrences': detailed_df['Number of Occurrences'].sum(),
        'Effort by Occurrence (Person-day)': detailed_df['Effort by Occurrence (Person-day)'].sum(),
        'Cost (FTE/Day)': '',  # Leave empty as it requires user input
        'Tech Debt ($) Effort x Cost': ''  # Will be calculated by Excel formula
    }
    excel_detailed_df = pd.concat([excel_detailed_df, pd.DataFrame([total_row])], ignore_index=True)
    
    # Add total row to summary sheet
    summary_total_row = {
        'Rule/Pattern': 'TOTAL',
        'Technology': '',
        'Number of Occurrences': summary_df['Number of Occurrences'].sum()
    }
    excel_summary_df = pd.concat([excel_summary_df, pd.DataFrame([summary_total_row])], ignore_index=True)
    
    # Save to Excel with two sheets
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        excel_detailed_df.to_excel(writer, index=False, sheet_name='Detailed Green Metrics')
        excel_summary_df.to_excel(writer, index=False, sheet_name='Summary by Rule')
    
    # Format Excel file
    wb = load_workbook(output_file)
    
    # Format detailed sheet
    ws_detail = wb['Detailed Green Metrics']
    
    # Add formulas for Tech Debt and Total
    for row in range(2, len(detailed_df) + 2):
        ws_detail[f'F{row}'] = f'=ROUND(D{row}*E{row}, 2)'
    
    # Format headers
    for cell in ws_detail[1]:
        cell.font = Font(bold=True)
    
    # Format numbers (2 decimal places)
    for row in ws_detail.iter_rows(min_row=2, max_row=len(detailed_df)+2, min_col=3, max_col=7):
        for cell in row:
            if cell.column_letter in ['C', 'D', 'F']:  # Numeric columns
                cell.number_format = numbers.FORMAT_NUMBER_00
    
    # Format total row
    total_row_num = len(detailed_df) + 2
    for cell in ws_detail[total_row_num]:
        if cell.column_letter in ['C', 'D']:
            cell.font = Font(bold=True)
    
    # Add formula for total Tech Debt
    ws_detail[f'F{total_row_num}'] = f'=SUM(F2:F{total_row_num-1})'
    
    # Format summary sheet
    ws_summary = wb['Summary by Rule']
    
    # Format headers
    for cell in ws_summary[1]:
        cell.font = Font(bold=True)
    
    # Format numbers
    for row in ws_summary.iter_rows(min_row=2, max_row=len(summary_df)+2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = numbers.FORMAT_NUMBER_00
    
    # Format total row
    summary_total_row_num = len(summary_df) + 2
    for cell in ws_summary[summary_total_row_num]:
        if cell.column_letter == 'C':
            cell.font = Font(bold=True)
    
    # Auto-size columns for both sheets
    for sheet in [ws_detail, ws_summary]:
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    wb.save(output_file)
    print(f"? Successfully saved data to {output_file}")
    print("?? Remember to enter cost rates in Column E to calculate Tech Debt")

def main():
    hl_instance, domain_id, application_id, api_key = load_config()
    if None in (hl_instance, domain_id, application_id, api_key):
        print("? Failed to load configuration")
        return
    
    detailed_df, summary_df = extract_green_data(hl_instance, domain_id, application_id, api_key)
    if detailed_df is not None and summary_df is not None:
        save_to_excel(detailed_df, summary_df, domain_id, application_id)
    else:
        print("? No data was extracted. Check the input parameters and API access.")

if __name__ == '__main__':
    main()