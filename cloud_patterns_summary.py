import requests
import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, numbers
from datetime import datetime
import os

def load_config():
    try:
        with open('config.json', 'r') as f:
            config = json.load(f)
            return (
                config.get('HLInstance'),
                config.get('domain_id'),
                config.get('api_key')
            )
    except Exception as e:
        print(f"Error loading config.json: {e}")
        return None, None, None

def get_domain_cloud_data(hl_instance, domain_id, api_key):
    url = f'https://{hl_instance}.casthighlight.com/WS2/cloud/requirements/{domain_id}'
    headers = {'Authorization': f'Bearer {api_key}'}
    try:
        response = requests.get(url, headers=headers, timeout=30)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"API request failed: {e}")
        return None

def extract_data(response_json):
    if not response_json or not isinstance(response_json, list):
        print("Unexpected or empty response format.")
        return None, None, None

    detailed_rows = []
    summary_by_pattern = {}
    pattern_app_map = {}
    all_apps_set = set()

    for item in response_json:
        pattern = item.get('display', 'N/A')
        technology = item.get('techno', {}).get('display', 'N/A')
        occurrences = item.get('roadBlocks', 0)
        effort_min = item.get('cloudEffort', 0)
        applications = item.get('applications', [])

        if not applications or occurrences == 0:
            continue

        effort_pd = round(effort_min / 480, 2)

        # Detailed sheet row
        detailed_rows.append([
            pattern, technology, occurrences, effort_pd, None, None
        ])

        # Summary by pattern with technology tracking
        if pattern not in summary_by_pattern:
            summary_by_pattern[pattern] = {'count': 0, 'technos': set()}
        summary_by_pattern[pattern]['count'] += occurrences
        summary_by_pattern[pattern]['technos'].add(technology)

        # Unique apps per pattern
        app_names = {app['name'] for app in applications}
        if pattern not in pattern_app_map:
            pattern_app_map[pattern] = set()
        pattern_app_map[pattern].update(app_names)

        # All apps
        all_apps_set.update(app_names)

    # Detailed Cloud Metrics
    detailed_df = pd.DataFrame(detailed_rows, columns=[
        'Rule/Pattern', 'Technology', 'Number of Occurrences',
        'Effort by Occurrence (Person-day)', 'Cost (FTE/Day)', 'Tech Debt ($) Effort x Cost'
    ]).sort_values(by='Number of Occurrences', ascending=False)

    # Summary by Rule – now includes Technologies column
    summary_df = pd.DataFrame([
        [pattern, ', '.join(sorted(data['technos'])), data['count']]
        for pattern, data in summary_by_pattern.items()
    ], columns=['Rule/Pattern', 'Technologies', 'Number of Occurrences']).sort_values(by='Number of Occurrences', ascending=False)

    # Unique Apps per Pattern
    unique_apps_df = pd.DataFrame([
        [pattern, len(apps)] for pattern, apps in pattern_app_map.items()
    ], columns=['Rule/Pattern', 'Unique Applications']).sort_values(by='Unique Applications', ascending=False)

    # Add total row to Unique Apps sheet
    unique_apps_df.loc[len(unique_apps_df.index)] = [
        'Unique apps across all patterns',
        len(all_apps_set)
    ]

    return detailed_df, summary_df, unique_apps_df

def save_to_excel(detailed_df, summary_df, unique_apps_df, domain_id):
    os.makedirs('output', exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f'output/domain_cloud_metrics_d{domain_id}_{timestamp}.xlsx'

    # Add TOTAL rows
    detailed_total = {
        'Rule/Pattern': 'TOTAL',
        'Technology': '',
        'Number of Occurrences': detailed_df['Number of Occurrences'].sum(),
        'Effort by Occurrence (Person-day)': detailed_df['Effort by Occurrence (Person-day)'].sum(),
        'Cost (FTE/Day)': '',
        'Tech Debt ($) Effort x Cost': ''
    }
    detailed_df = pd.concat([detailed_df, pd.DataFrame([detailed_total])], ignore_index=True)

    summary_total = {
        'Rule/Pattern': 'TOTAL',
        'Technologies': '',
        'Number of Occurrences': summary_df['Number of Occurrences'].sum()
    }
    summary_df = pd.concat([summary_df, pd.DataFrame([summary_total])], ignore_index=True)

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        detailed_df.to_excel(writer, index=False, sheet_name='Detailed Cloud Metrics')
        summary_df.to_excel(writer, index=False, sheet_name='Summary by Rule')
        unique_apps_df.to_excel(writer, index=False, sheet_name='Pattern in Unique Apps')

    # Formatting
    wb = load_workbook(output_file)

    def format_sheet(ws, last_data_row, numeric_cols, bold_last=False):
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in ws.iter_rows(min_row=2, max_row=last_data_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.column_letter in numeric_cols:
                    cell.number_format = numbers.FORMAT_NUMBER_00
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max_length + 2

        if bold_last:
            for cell in ws[last_data_row + 1]:
                cell.font = Font(bold=True)

    ws_detail = wb['Detailed Cloud Metrics']
    for row in range(2, len(detailed_df) + 1):
        ws_detail[f'F{row}'] = f'=ROUND(D{row}*E{row}, 2)'
    format_sheet(ws_detail, len(detailed_df), numeric_cols=['C', 'D', 'F'], bold_last=True)

    ws_summary = wb['Summary by Rule']
    format_sheet(ws_summary, len(summary_df), numeric_cols=['C'], bold_last=True)

    ws_unique = wb['Pattern in Unique Apps']
    format_sheet(ws_unique, len(unique_apps_df), numeric_cols=['B'], bold_last=True)

    wb.save(output_file)
    print(f"✅ Data saved to {output_file}")
    print("ℹ️ Enter cost rates in Column E to calculate Tech Debt.")

def main():
    hl_instance, domain_id, api_key = load_config()
    if None in (hl_instance, domain_id, api_key):
        print("❌ Failed to load configuration")
        return

    response_json = get_domain_cloud_data(hl_instance, domain_id, api_key)
    detailed_df, summary_df, unique_apps_df = extract_data(response_json)

    if detailed_df is not None:
        save_to_excel(detailed_df, summary_df, unique_apps_df, domain_id)
    else:
        print("❌ No data extracted. Please check the API response.")

if __name__ == '__main__':
    main()
